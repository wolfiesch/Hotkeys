#!/usr/bin/env python3
"""
Create a styled Excel file for ExcelDatabookLayers.ahk hotkey reference
Prioritizes readability and quick hotkey lookup
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def create_styled_hotkeys_excel():
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hotkey Reference"

    # Define color themes for each category
    colors = {
        "PowerPoint": {"bg": "E6E6FA", "text": "4B0082"},  # Light purple, dark purple
        "PASTE": {"bg": "E6F3FF", "text": "0066CC"},  # Light blue, dark blue
        "FORMAT": {"bg": "E6FFE6", "text": "006600"},  # Light green, dark green
        "ALIGNMENT": {"bg": "FFE6CC", "text": "CC6600"},  # Light orange, dark orange
        "BORDERS": {"bg": "FFE6E6", "text": "CC0000"},  # Light red, dark red
        "SIZING": {"bg": "E6FFFF", "text": "006666"},  # Light teal, dark teal
        "NAVIGATION": {"bg": "FFFACD", "text": "B8860B"},  # Light yellow, dark yellow
        "DATA": {"bg": "FFE6F3", "text": "CC0066"},  # Light pink, dark pink
        "CLEAR": {"bg": "F0F0F0", "text": "666666"},  # Light gray, dark gray
    }

    # Headers
    headers = [
        "Category",
        "Hotkey Display",
        "Key Combination",
        "Function",
        "Description",
        "Line #",
    ]

    # Hotkey data (extracted from the AHK file)
    hotkeys = [
        # PowerPoint
        [
            "PowerPoint",
            "Ctrl+Alt+Shift+S",
            "^!+s",
            "PowerPoint Sequence",
            "Send Tab Tab 70 Enter Tab×4 4.57 Tab×2 21.47",
            "30",
        ],
        [
            "PowerPoint",
            "Ctrl+Alt+Shift+F",
            "^!+f",
            "Format Object Pane",
            "Alt+4 wait Ctrl+A wait 70 Enter",
            "52",
        ],
        # PASTE
        [
            "PASTE",
            "CapsLock+V",
            "SC03A & v",
            "Paste Values",
            "Paste values only without formulas or formatting",
            "756",
        ],
        [
            "PASTE",
            "CapsLock+F",
            "SC03A & f",
            "Paste Formulas",
            "Paste formulas only",
            "757",
        ],
        [
            "PASTE",
            "Ctrl+CapsLock+F",
            "Ctrl+SC03A & f",
            "Freeze Panes",
            "Freeze panes at active cell",
            "757",
        ],
        [
            "PASTE",
            "CapsLock+T",
            "SC03A & t",
            "Paste Formats",
            "Paste formatting only",
            "765",
        ],
        [
            "PASTE",
            "CapsLock+W",
            "SC03A & w",
            "Paste Column Widths",
            "Paste column width information",
            "766",
        ],
        [
            "PASTE",
            "CapsLock+S",
            "SC03A & s",
            "Paste Formulas + Format",
            "Paste both formulas and formatting",
            "767",
        ],
        [
            "PASTE",
            "CapsLock+X",
            "SC03A & x",
            "Paste Values (Transpose)",
            "Paste values with transpose operation",
            "768",
        ],
        [
            "PASTE",
            "CapsLock+L",
            "SC03A & l",
            "Paste Link",
            "Create linked paste",
            "769",
        ],
        [
            "PASTE",
            "CapsLock+P",
            "SC03A & p",
            "Paste Special Dialog",
            "Open Paste Special dialog box",
            "770",
        ],
        [
            "PASTE",
            "CapsLock+Numpad/",
            "SC03A & NumpadDiv",
            "Toggle Filter",
            "Toggle AutoFilter on/off",
            "773",
        ],
        [
            "PASTE",
            "CapsLock+Numpad*",
            "SC03A & NumpadMult",
            "Clear Filter",
            "Clear all filters",
            "774",
        ],
        [
            "PASTE",
            "CapsLock+Numpad+",
            "SC03A & NumpadAdd",
            "Paste Add",
            "Paste operation with addition",
            "775",
        ],
        [
            "PASTE",
            "CapsLock+Numpad-",
            "SC03A & NumpadSub",
            "Paste Subtract",
            "Paste operation with subtraction",
            "776",
        ],
        # FORMAT
        [
            "FORMAT",
            "CapsLock+1",
            "SC03A & 1",
            "Custom Font Color",
            "Run custom font color macro",
            "779",
        ],
        [
            "FORMAT",
            "CapsLock+2",
            "SC03A & 2",
            "Subtotal Format",
            "Apply subtotal row formatting",
            "780",
        ],
        [
            "FORMAT",
            "CapsLock+3",
            "SC03A & 3",
            "Major Total Format",
            "Apply major total row formatting",
            "781",
        ],
        [
            "FORMAT",
            "CapsLock+4",
            "SC03A & 4",
            "Grand Total Format",
            "Apply grand total row formatting",
            "782",
        ],
        [
            "FORMAT",
            "CapsLock+6",
            "SC03A & 6",
            "Set Row Height 5pt",
            "Set row height to 5 points",
            "783",
        ],
        [
            "FORMAT",
            "CapsLock+9",
            "SC03A & 9",
            "General Format",
            "Apply general number format",
            "786",
        ],
        [
            "FORMAT",
            "CapsLock+A",
            "SC03A & a",
            "Set Row Height 5pt",
            "Set row height to 5 points",
            "787",
        ],
        [
            "FORMAT",
            "CapsLock+K",
            "SC03A & k",
            "Custom Number Format",
            "Run custom number format macro",
            "788",
        ],
        [
            "FORMAT",
            "CapsLock+5",
            "SC03A & 5",
            "Percent Format",
            "Apply percentage number format",
            "789",
        ],
        [
            "FORMAT",
            "CapsLock+M",
            "SC03A & m",
            "Month Format",
            "Apply month format (mmm-yy)",
            "790",
        ],
        [
            "FORMAT",
            "CapsLock+D",
            "SC03A & d",
            "Date Format",
            "Apply date number format",
            "791",
        ],
        # ALIGNMENT
        [
            "ALIGNMENT",
            "CapsLock+F1",
            "SC03A & F1",
            "Align Left",
            "Set text alignment to left",
            "794",
        ],
        [
            "ALIGNMENT",
            "CapsLock+F2",
            "SC03A & F2",
            "Align Center",
            "Set text alignment to center",
            "795",
        ],
        [
            "ALIGNMENT",
            "CapsLock+F3",
            "SC03A & F3",
            "Align Right",
            "Set text alignment to right",
            "796",
        ],
        [
            "ALIGNMENT",
            "CapsLock+F4",
            "SC03A & F4",
            "Toggle Wrap Text",
            "Toggle text wrapping in cells",
            "797",
        ],
        # BORDERS
        [
            "BORDERS",
            "CapsLock+R",
            "SC03A & r",
            "Right Border",
            "Apply right border macro",
            "800",
        ],
        [
            "SIZING",
            "Ctrl+CapsLock+R",
            "Ctrl+SC03A & r",
            "AutoFit Rows",
            "Auto-adjust row heights to fit content",
            "800",
        ],
        [
            "BORDERS",
            "CapsLock+B",
            "SC03A & b",
            "Bottom Border",
            "Apply bottom thin border macro",
            "808",
        ],
        [
            "BORDERS",
            "CapsLock+O",
            "SC03A & o",
            "Outline Borders",
            "Apply outline borders to selection",
            "809",
        ],
        [
            "BORDERS",
            "CapsLock+I",
            "SC03A & i",
            "Inside Borders",
            "Apply inside borders to selection",
            "810",
        ],
        [
            "BORDERS",
            "CapsLock+C",
            "SC03A & c",
            "Clear Borders",
            "Clear all borders",
            "811",
        ],
        [
            "SIZING",
            "Ctrl+CapsLock+C",
            "Ctrl+SC03A & c",
            "AutoFit Columns",
            "Auto-adjust column widths to fit content",
            "811",
        ],
        [
            "BORDERS",
            "CapsLock+Y",
            "SC03A & y",
            "Top Double Border",
            "Apply double border to top",
            "819",
        ],
        [
            "BORDERS",
            "CapsLock+J",
            "SC03A & j",
            "Left Thick Border",
            "Apply thick border to left",
            "820",
        ],
        [
            "BORDERS",
            "CapsLock+;",
            "SC03A & ;",
            "Right Thick Border",
            "Apply thick border to right",
            "821",
        ],
        # SIZING
        [
            "SIZING",
            "CapsLock+Q",
            "SC03A & q",
            "Column Width 0.5",
            "Set column width to 0.5",
            "824",
        ],
        [
            "SIZING",
            "Ctrl+CapsLock+Q",
            "Ctrl+SC03A & q",
            "Row Height 5pt",
            "Set row height to 5 points",
            "824",
        ],
        [
            "SIZING",
            "CapsLock+F5",
            "SC03A & F5",
            "AutoFit Columns",
            "Auto-adjust column widths to fit content",
            "832",
        ],
        [
            "SIZING",
            "CapsLock+F6",
            "SC03A & F6",
            "AutoFit Rows",
            "Auto-adjust row heights to fit content",
            "833",
        ],
        [
            "SIZING",
            "CapsLock+F11",
            "SC03A & F11",
            "Increase Indent",
            "Increase text indentation",
            "834",
        ],
        [
            "SIZING",
            "CapsLock+F12",
            "SC03A & F12",
            "Decrease Indent",
            "Decrease text indentation",
            "835",
        ],
        [
            "SIZING",
            "CapsLock+Numpad.",
            "SC03A & NumpadDot",
            "Add Decimal Place",
            "Increase decimal places in number format",
            "836",
        ],
        [
            "SIZING",
            "CapsLock+Numpad0",
            "SC03A & Numpad0",
            "Remove Decimal Place",
            "Decrease decimal places in number format",
            "837",
        ],
        # NAVIGATION
        [
            "NAVIGATION",
            "CapsLock+[",
            "SC03A & [",
            "Previous Divider",
            "Jump to previous data boundary",
            "840",
        ],
        [
            "NAVIGATION",
            "CapsLock+]",
            "SC03A & ]",
            "Next Divider",
            "Jump to next data boundary",
            "841",
        ],
        [
            "NAVIGATION",
            "CapsLock+=",
            "SC03A & =",
            "First Block Edge",
            "Jump to first block edge",
            "842",
        ],
        [
            "NAVIGATION",
            "CapsLock+-",
            "SC03A & -",
            "Last Block Edge",
            "Jump to last block edge",
            "843",
        ],
        [
            "NAVIGATION",
            "CapsLock+,",
            "SC03A & ,",
            "Previous Sheet",
            "Switch to previous worksheet",
            "844",
        ],
        [
            "NAVIGATION",
            "CapsLock+.",
            "SC03A & .",
            "Next Sheet",
            "Switch to next worksheet",
            "845",
        ],
        [
            "NAVIGATION",
            "CapsLock+G",
            "SC03A & g",
            "Go To",
            "Open Go To dialog (Ctrl+G)",
            "846",
        ],
        [
            "NAVIGATION",
            "CapsLock+8",
            "SC03A & 8",
            "Select Current Region",
            "Select current data region (Ctrl+Shift+8)",
            "847",
        ],
        [
            "NAVIGATION",
            "CapsLock+H",
            "SC03A & h",
            "Custom Macro H",
            "Run custom macro Ctrl+Alt+Shift+H",
            "848",
        ],
        [
            "NAVIGATION",
            "CapsLock+Numpad8",
            "SC03A & Numpad8",
            "Ctrl+Up",
            "Move to edge of data block upward",
            "851",
        ],
        [
            "NAVIGATION",
            "CapsLock+Numpad2",
            "SC03A & Numpad2",
            "Ctrl+Down",
            "Move to edge of data block downward",
            "852",
        ],
        [
            "NAVIGATION",
            "CapsLock+Numpad4",
            "SC03A & Numpad4",
            "Ctrl+Left",
            "Move to edge of data block leftward",
            "853",
        ],
        [
            "NAVIGATION",
            "CapsLock+Numpad6",
            "SC03A & Numpad6",
            "Ctrl+Right",
            "Move to edge of data block rightward",
            "854",
        ],
        [
            "NAVIGATION",
            "CapsLock+Numpad7",
            "SC03A & Numpad7",
            "Ctrl+Home",
            "Move to cell A1",
            "855",
        ],
        [
            "NAVIGATION",
            "CapsLock+Numpad9",
            "SC03A & Numpad9",
            "Ctrl+End",
            "Move to last used cell",
            "856",
        ],
        [
            "NAVIGATION",
            "CapsLock+Space",
            "SC03A & Space",
            "Show Help",
            "Display hotkey reference overlay",
            "872",
        ],
        # DATA
        [
            "DATA",
            "CapsLock+U",
            "SC03A & u",
            "Trim In Place",
            "Remove leading/trailing spaces from text",
            "859",
        ],
        [
            "DATA",
            "CapsLock+F8",
            "SC03A & F8",
            "Clean In Place",
            "Remove non-printable characters",
            "860",
        ],
        [
            "DATA",
            "CapsLock+N",
            "SC03A & n",
            "Convert to Number",
            "Convert text to numbers using Text to Columns",
            "861",
        ],
        [
            "DATA",
            "CapsLock+E",
            "SC03A & e",
            "Text to Columns",
            "Open Text to Columns wizard",
            "862",
        ],
        [
            "DATA",
            "CapsLock+F7",
            "SC03A & F7",
            "Toggle AutoFilter",
            "Toggle AutoFilter on/off (Ctrl+Shift+L)",
            "863",
        ],
        # CLEAR
        [
            "CLEAR",
            "CapsLock+Shift+Delete",
            "SC03A & +Delete",
            "Clear Formats",
            "Clear cell formatting only",
            "867",
        ],
        [
            "CLEAR",
            "CapsLock+Delete",
            "SC03A & Delete",
            "Clear Contents",
            "Clear cell contents only",
            "868",
        ],
        [
            "CLEAR",
            "CapsLock+Ctrl+Delete",
            "SC03A & ^Delete",
            "Clear All",
            "Clear both contents and formatting",
            "869",
        ],
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="2F4F4F", end_color="2F4F4F", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thick"),
        )

    # Write data with styling
    current_category = None
    row_num = 2

    for hotkey in hotkeys:
        category = hotkey[0]

        # Add category separator if new category
        if category != current_category:
            current_category = category

            # Add a subtle separator row if not the first category
            if row_num > 2:
                row_num += 1

        # Write data
        for col, value in enumerate(hotkey, 1):
            cell = ws.cell(row=row_num, column=col, value=value)

            # Category styling
            if col == 1:  # Category column
                cell.font = Font(bold=True, size=10, color=colors[category]["text"])
                cell.fill = PatternFill(
                    start_color=colors[category]["bg"],
                    end_color=colors[category]["bg"],
                    fill_type="solid",
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Hotkey display styling
            elif col == 2:  # Hotkey display column
                cell.font = Font(bold=True, size=11, color="000080")
                cell.fill = PatternFill(
                    start_color="F8F8FF", end_color="F8F8FF", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Function name styling
            elif col == 4:  # Function column
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Description styling
            elif col == 5:  # Description column
                cell.font = Font(size=9)
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

            # Line number styling
            elif col == 6:  # Line number column
                cell.font = Font(size=9, color="666666")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            else:
                cell.font = Font(size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Add borders
            cell.border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )

        row_num += 1

    # Set column widths for readability
    column_widths = {
        "A": 12,  # Category
        "B": 25,  # Hotkey Display
        "C": 20,  # Key Combination
        "D": 22,  # Function
        "E": 45,  # Description
        "F": 8,  # Line #
    }

    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    # Set row heights
    for row in range(2, row_num):
        ws.row_dimensions[row].height = 20

    # Header row height
    ws.row_dimensions[1].height = 25

    # Freeze panes (keep header visible)
    ws.freeze_panes = "A2"

    # Add AutoFilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_num-1}"

    # Add title and metadata to a separate sheet
    info_ws = wb.create_sheet(title="Info")
    info_ws["A1"] = "ExcelDatabookLayers.ahk Hotkey Reference"
    info_ws["A1"].font = Font(bold=True, size=16)
    info_ws["A3"] = "Generated from ExcelDatabookLayers.ahk"
    info_ws["A4"] = f"Total Hotkeys: {len(hotkeys)}"
    info_ws["A5"] = "Categories: " + ", ".join(colors.keys())
    info_ws.column_dimensions["A"].width = 40

    # Save the workbook
    wb.save(r"C:\Users\wschoenberger\Hotkeys\ExcelDatabookLayers_Hotkeys.xlsx")
    print("Excel file created successfully: ExcelDatabookLayers_Hotkeys.xlsx")
    print(f"Total hotkeys documented: {len(hotkeys)}")
    print("Categories:", list(colors.keys()))


if __name__ == "__main__":
    create_styled_hotkeys_excel()
